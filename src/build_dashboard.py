"""
Generate region-specific HTML dashboards (Switzerland, UK, …).

Each dashboard includes:
  • KPI grid (current salary, real values, inflation, gain/loss)
  • Interactive calculator (visitors plug in their own numbers)
  • Charts (nominal vs real, hourly, inflation trend, gain/loss)
  • Year-by-year table
  • Top navigation between regions
  • Mobile-first responsive layout with refined breakpoints
"""
from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import pandas as pd

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Formatting
# ---------------------------------------------------------------------------

def _fmt_currency(value, region: dict, decimals: int = 0) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "—"
    sym = region["currency_symbol"]
    sep = region["thousand_sep"]
    s = f"{abs(value):,.{decimals}f}".replace(",", sep)
    sign = "−" if value < 0 else ""
    if sym == "£" or sym == "$":
        return f"{sign}{sym}{s}"
    return f"{sign}{sym} {s}"


def _fmt_pct(v, decimals: int = 1) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    sign = "+" if v >= 0 else ""
    return f"{sign}{v:.{decimals}f}%"


# ---------------------------------------------------------------------------
# Chart payloads
# ---------------------------------------------------------------------------

def _chart_yearly(df: pd.DataFrame, region: dict) -> str:
    years = df["year"].tolist()
    nom = [round(v, 0) if pd.notna(v) else None for v in df["nominal_annual"]]
    real = [round(v, 0) if pd.notna(v) else None for v in df["real_annual"]]
    return json.dumps({
        "traces": [
            {"x": years, "y": nom,  "name": "Nominal", "type": "bar", "marker": {"color": "#3b82f6"}},
            {"x": years, "y": real, "name": f"Real ({region['base_year']} {region['currency']})", "type": "bar", "marker": {"color": "#22c55e"}},
        ],
        "layout": {
            "title": f"Annual salary: nominal vs real ({region['currency']})",
            "barmode": "group",
            "yaxis": {"title": region["currency"], "tickformat": ",.0f"},
            "xaxis": {"title": "Year", "type": "category"},
            "legend": {"orientation": "h", "y": -0.2},
            "margin": {"t": 50, "b": 60, "l": 70, "r": 20},
        },
    })


def _chart_hourly(df: pd.DataFrame, region: dict) -> str:
    years = df["year"].tolist()
    nom = [round(v, 2) if pd.notna(v) else None for v in df["nominal_hourly"]]
    real = [round(v, 2) if pd.notna(v) else None for v in df["real_hourly"]]
    return json.dumps({
        "traces": [
            {"x": years, "y": nom,  "name": "Nominal hourly", "type": "scatter",
             "mode": "lines+markers", "line": {"color": "#3b82f6", "width": 3}, "marker": {"size": 8}},
            {"x": years, "y": real, "name": f"Real hourly ({region['base_year']} {region['currency']})",
             "type": "scatter", "mode": "lines+markers", "line": {"color": "#22c55e", "width": 3}, "marker": {"size": 8}},
        ],
        "layout": {
            "title": f"Hourly rate: nominal vs real ({region['currency']}/hour)",
            "yaxis": {"title": f"{region['currency']}/hour", "tickformat": ",.2f"},
            "xaxis": {"title": "Year", "type": "category"},
            "legend": {"orientation": "h", "y": -0.2},
            "margin": {"t": 50, "b": 60, "l": 70, "r": 20},
        },
    })


def _chart_inflation(df_monthly: pd.DataFrame, region: dict) -> str:
    if df_monthly.empty:
        return "null"
    periods = df_monthly["period"].tolist()
    vals = [round(v, 2) for v in df_monthly["cumulative_inflation_pct"]]
    return json.dumps({
        "traces": [{
            "x": periods, "y": vals, "name": f"Inflation since {region['base_year']}",
            "type": "scatter", "mode": "lines", "fill": "tozeroy",
            "line": {"color": region["theme_color"], "width": 2},
            "fillcolor": "rgba(245,158,11,0.10)",
        }],
        "layout": {
            "title": f"Cumulative CPI inflation since {region['base_year']} (%)",
            "yaxis": {"title": f"% vs {region['base_year']} avg", "ticksuffix": "%"},
            "xaxis": {"title": "Month"},
            "margin": {"t": 50, "b": 60, "l": 70, "r": 20},
            "shapes": [{"type": "line", "x0": periods[0], "x1": periods[-1],
                        "y0": 0, "y1": 0, "line": {"dash": "dot", "color": "#94a3b8"}}],
        },
    })


def _chart_gain(df_monthly: pd.DataFrame, region: dict) -> str:
    if df_monthly.empty:
        return "null"
    periods = df_monthly["period"].tolist()
    vals = [round(v, 0) for v in df_monthly["real_gain_loss_annual"]]
    colors = ["#16a34a" if v >= 0 else "#dc2626" for v in vals]
    return json.dumps({
        "traces": [{
            "x": periods, "y": vals, "name": "Annual gain/loss vs base",
            "type": "bar", "marker": {"color": colors},
        }],
        "layout": {
            "title": f"Real annual gain / loss vs {region['base_year']} ({region['currency']})",
            "yaxis": {"title": region["currency"], "tickformat": ",.0f"},
            "xaxis": {"title": "Month"},
            "margin": {"t": 50, "b": 60, "l": 70, "r": 20},
            "shapes": [{"type": "line", "x0": periods[0], "x1": periods[-1],
                        "y0": 0, "y1": 0, "line": {"dash": "dot", "color": "#94a3b8"}}],
        },
    })


# ---------------------------------------------------------------------------
# Table
# ---------------------------------------------------------------------------

def _table_rows(df: pd.DataFrame, region: dict) -> str:
    out = []
    for _, r in df.iterrows():
        cls = "positive" if (r["annual_gain_loss_vs_base"] or 0) >= 0 else "negative"
        out.append(f"""
        <tr>
          <td class="bold">{int(r['year'])}</td>
          <td>{_fmt_currency(r['nominal_monthly'], region)}</td>
          <td>{_fmt_currency(r['nominal_annual'], region)}</td>
          <td>{_fmt_pct(r['nominal_pct_vs_base'])}</td>
          <td>{_fmt_pct(r['cumulative_inflation_pct'])}</td>
          <td>{_fmt_currency(r['real_monthly'], region)}</td>
          <td>{_fmt_currency(r['real_annual'], region)}</td>
          <td class="{cls}">{_fmt_currency(r['annual_gain_loss_vs_base'], region)}</td>
          <td class="{cls}">{_fmt_pct(r['real_pct_vs_base'])}</td>
          <td>{_fmt_currency(r['nominal_hourly'], region, 2)}</td>
          <td>{_fmt_currency(r['real_hourly'], region, 2)}</td>
        </tr>""")
    return "\n".join(out)


# ---------------------------------------------------------------------------
# Top navigation
# ---------------------------------------------------------------------------

def _build_nav(active_code: str, all_regions: Iterable[dict]) -> str:
    items = []
    for r in all_regions:
        href = "index.html" if r["code"] == "ch" else f"{r['code']}.html"
        cls = "nav-item active" if r["code"] == active_code else "nav-item"
        items.append(f'<a class="{cls}" href="{href}">{r["flag"]} {r["name"]}</a>')
    return f'<nav class="region-nav">{"".join(items)}</nav>'


# ---------------------------------------------------------------------------
# Main HTML
# ---------------------------------------------------------------------------

def build_html(
    summary: dict,
    df_yearly: pd.DataFrame,
    df_monthly: pd.DataFrame,
    df_cpi_full: pd.DataFrame,
    region: dict,
    all_regions: Iterable[dict],
) -> str:
    generated = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
    base_year = summary["base_year"]
    nav = _build_nav(region["code"], all_regions)

    nom_cls = "positive" if summary["nominal_increase_vs_base_pct"] >= 0 else "negative"
    real_cls = "positive" if summary["real_increase_vs_base_pct"] >= 0 else "negative"
    gl_a_cls = "positive" if summary["annual_gain_loss_vs_base"] >= 0 else "negative"
    gl_m_cls = "positive" if summary["monthly_gain_loss_vs_base"] >= 0 else "negative"

    chart_yearly = _chart_yearly(df_yearly, region)
    chart_hourly = _chart_hourly(df_yearly, region)
    chart_inflation = _chart_inflation(df_monthly, region)
    chart_gain = _chart_gain(df_monthly, region)
    table_rows = _table_rows(df_yearly, region)

    # Calculator data: pass FULL CPI history (for compare-two-years feature)
    cpi_periods = df_cpi_full[["period", "cpi"]].to_dict(orient="records")
    available_years = sorted({int(p["period"][:4]) for p in cpi_periods})

    salary_history_defaults = [
        {"year": int(r["year"]), "annual": int(r["nominal_annual"])}
        for _, r in df_yearly.iterrows()
    ]

    calc_payload = json.dumps({
        "cpi_data": cpi_periods,
        "available_years": available_years,
        "default_base_year": base_year,
        "default_weekly_hours": region["weekly_hours"],
        "default_weeks": region["weeks_per_year"],
        "currency_symbol": region["currency_symbol"],
        "thousand_sep": region["thousand_sep"],
        "salary_history_defaults": salary_history_defaults,
    })

    fmt_cur = lambda v, d=0: _fmt_currency(v, region, d)
    fmt_pct = _fmt_pct

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
  <meta name="theme-color" content="{region['theme_color']}">
  <title>{region['flag']} Salary Tracker · {region['name']}</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
  <script src="https://cdn.plot.ly/plotly-2.27.0.min.js" charset="utf-8"></script>
  <style>
    :root {{
      --bg: #f8fafc;
      --card: #ffffff;
      --primary: {region['theme_color']};
      --primary-dark: #0f172a;
      --accent: #3b82f6;
      --positive: #16a34a;
      --negative: #dc2626;
      --text: #0f172a;
      --muted: #64748b;
      --muted-2: #94a3b8;
      --border: #e2e8f0;
      --border-strong: #cbd5e1;
      --radius: 12px;
      --shadow-sm: 0 1px 2px rgba(15,23,42,.04);
      --shadow: 0 1px 3px rgba(15,23,42,.06), 0 1px 2px rgba(15,23,42,.04);
      --shadow-lg: 0 10px 25px -5px rgba(15,23,42,.08), 0 8px 10px -6px rgba(15,23,42,.05);
    }}
    @media (prefers-color-scheme: dark) {{
      :root {{
        --bg: #0b1220;
        --card: #111827;
        --primary-dark: #e2e8f0;
        --text: #e2e8f0;
        --muted: #94a3b8;
        --muted-2: #64748b;
        --border: #1f2937;
        --border-strong: #374151;
        --shadow-sm: 0 1px 2px rgba(0,0,0,.3);
        --shadow: 0 1px 3px rgba(0,0,0,.5);
        --shadow-lg: 0 10px 25px -5px rgba(0,0,0,.5);
      }}
    }}

    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    html {{ scroll-behavior: smooth; }}
    body {{
      font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
      background: var(--bg);
      color: var(--text);
      font-size: 15px;
      line-height: 1.5;
      -webkit-font-smoothing: antialiased;
      text-rendering: optimizeLegibility;
    }}

    /* ────────── Header ────────── */
    header {{
      background: linear-gradient(135deg, var(--primary) 0%, color-mix(in srgb, var(--primary) 70%, #000) 100%);
      color: white;
      padding: 1.25rem 1.25rem 1rem;
      box-shadow: var(--shadow);
    }}
    .header-inner {{ max-width: 1280px; margin: 0 auto; display: flex; align-items: center; gap: 1rem; flex-wrap: wrap; }}
    .header-text h1 {{ font-size: 1.25rem; font-weight: 700; letter-spacing: -0.02em; }}
    .header-text p {{ font-size: 0.82rem; opacity: 0.85; margin-top: 2px; }}
    .header-flag {{ font-size: 2rem; line-height: 1; }}

    /* ────────── Region nav ────────── */
    .region-nav {{
      display: flex; gap: 0.5rem; flex-wrap: wrap;
      max-width: 1280px; margin: 0 auto; padding: 0.75rem 1.25rem;
      background: var(--card); border-bottom: 1px solid var(--border);
    }}
    .nav-item {{
      padding: 0.5rem 0.9rem; border-radius: 999px; font-size: 0.85rem;
      font-weight: 500; color: var(--muted); text-decoration: none;
      border: 1px solid var(--border); transition: all 0.15s ease;
      white-space: nowrap;
    }}
    .nav-item:hover {{ color: var(--text); border-color: var(--border-strong); }}
    .nav-item.active {{ background: var(--primary); color: white; border-color: var(--primary); }}

    .container {{ max-width: 1280px; margin: 0 auto; padding: 1.25rem; }}

    /* ────────── Section titles ────────── */
    .section-title {{
      font-size: 0.7rem; font-weight: 700;
      letter-spacing: 0.1em; text-transform: uppercase;
      color: var(--muted); margin: 1.75rem 0 0.85rem;
    }}
    .section-title:first-child {{ margin-top: 0.5rem; }}

    /* ────────── KPI cards ────────── */
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(180px, 1fr));
      gap: 0.75rem;
    }}
    .kpi-card {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: var(--radius);
      padding: 1rem 1.1rem;
      box-shadow: var(--shadow-sm);
      transition: transform 0.15s ease, box-shadow 0.15s ease;
    }}
    .kpi-card:hover {{ transform: translateY(-1px); box-shadow: var(--shadow); }}
    .kpi-label {{
      font-size: 0.7rem; font-weight: 600; color: var(--muted);
      text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 0.4rem;
    }}
    .kpi-value {{
      font-size: 1.4rem; font-weight: 700;
      color: var(--text); word-break: break-word;
      letter-spacing: -0.02em;
    }}
    .kpi-value.positive {{ color: var(--positive); }}
    .kpi-value.negative {{ color: var(--negative); }}
    .kpi-sub {{ font-size: 0.75rem; color: var(--muted-2); margin-top: 0.3rem; }}
    .positive {{ color: var(--positive) !important; }}
    .negative {{ color: var(--negative) !important; }}

    /* ────────── Calculator ────────── */
    .calculator {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: var(--radius);
      padding: 1.25rem;
      box-shadow: var(--shadow);
    }}
    .calc-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 1.25rem;
    }}
    @media (max-width: 720px) {{ .calc-grid {{ grid-template-columns: 1fr; }} }}
    .calc-inputs {{ display: flex; flex-direction: column; gap: 0.85rem; }}
    .calc-field {{ display: flex; flex-direction: column; gap: 0.35rem; }}
    .calc-field label {{
      font-size: 0.75rem; font-weight: 600;
      color: var(--muted); text-transform: uppercase; letter-spacing: 0.05em;
    }}
    .calc-field input, .calc-field select {{
      width: 100%; padding: 0.6rem 0.8rem; font-size: 1rem;
      font-family: inherit;
      border: 1.5px solid var(--border-strong); border-radius: 8px;
      background: var(--bg); color: var(--text);
      transition: border-color 0.15s ease;
    }}
    .calc-field input:focus, .calc-field select:focus {{
      outline: none; border-color: var(--accent);
    }}
    .calc-results {{
      display: grid; grid-template-columns: 1fr 1fr; gap: 0.6rem;
      align-content: start;
    }}
    .calc-result {{
      background: var(--bg); border-radius: 10px;
      padding: 0.75rem 0.9rem; border: 1px solid var(--border);
    }}
    .calc-result .lbl {{
      font-size: 0.7rem; font-weight: 600; color: var(--muted);
      text-transform: uppercase; letter-spacing: 0.05em;
    }}
    .calc-result .val {{
      font-size: 1.05rem; font-weight: 700; color: var(--text);
      margin-top: 0.25rem; letter-spacing: -0.01em;
    }}
    .calc-result.full {{ grid-column: 1 / -1; }}
    .calc-result.headline {{
      background: linear-gradient(135deg, color-mix(in srgb, var(--accent) 12%, var(--card)),
                                          color-mix(in srgb, var(--accent) 4%, var(--card)));
      border-color: color-mix(in srgb, var(--accent) 30%, var(--border));
    }}
    .calc-result.headline .val {{ font-size: 1.35rem; }}
    .calc-tabs {{
      display: flex; gap: 0.4rem; margin-bottom: 1rem;
      border-bottom: 1px solid var(--border);
      padding-bottom: 0;
    }}
    .calc-tab {{
      background: none; border: none; cursor: pointer;
      padding: 0.6rem 1rem; font-size: 0.9rem; font-weight: 600;
      color: var(--muted); font-family: inherit;
      border-bottom: 2.5px solid transparent;
      margin-bottom: -1px;
      transition: color 0.15s ease, border-color 0.15s ease;
    }}
    .calc-tab:hover {{ color: var(--text); }}
    .calc-tab.active {{ color: var(--accent); border-bottom-color: var(--accent); }}
    .calc-pane {{ display: none; }}
    .calc-pane.active {{ display: block; animation: fadeIn 0.2s ease; }}
    @keyframes fadeIn {{ from {{ opacity: 0; transform: translateY(4px); }} to {{ opacity: 1; transform: none; }} }}
    .calc-hint {{
      font-size: 0.85rem; color: var(--muted); margin-bottom: 1rem; line-height: 1.6;
    }}
    .calc-hint em {{ font-style: normal; font-weight: 600; color: var(--text); }}
    .calc-row {{ display: grid; grid-template-columns: 1fr 1.4fr; gap: 0.6rem; }}

    /* ────────── Charts ────────── */
    .chart-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(min(100%, 460px), 1fr));
      gap: 1rem;
    }}
    .chart-card {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: var(--radius);
      padding: 0.75rem;
      box-shadow: var(--shadow-sm);
    }}
    .chart-card .plotly-graph-div {{ width: 100% !important; }}

    /* ────────── Table ────────── */
    .table-card {{
      background: var(--card); border: 1px solid var(--border);
      border-radius: var(--radius); overflow: hidden;
      box-shadow: var(--shadow-sm);
    }}
    .table-scroll {{ overflow-x: auto; -webkit-overflow-scrolling: touch; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 0.83rem; }}
    thead {{ background: var(--primary-dark); color: white; }}
    @media (prefers-color-scheme: dark) {{
      thead {{ background: var(--border-strong); color: var(--text); }}
    }}
    th {{
      padding: 0.7rem 0.8rem; text-align: right; font-weight: 600;
      white-space: nowrap; font-size: 0.72rem; letter-spacing: 0.02em;
      text-transform: uppercase;
    }}
    th:first-child {{ text-align: left; }}
    td {{
      padding: 0.6rem 0.8rem; text-align: right;
      border-top: 1px solid var(--border); white-space: nowrap;
      font-variant-numeric: tabular-nums;
    }}
    td:first-child {{ text-align: left; }}
    tbody tr {{ transition: background 0.1s ease; }}
    tbody tr:hover {{ background: color-mix(in srgb, var(--accent) 6%, transparent); }}
    .bold {{ font-weight: 600; }}

    /* ────────── Explainer & footer ────────── */
    .explainer {{
      background: color-mix(in srgb, var(--accent) 8%, var(--card));
      border-left: 3px solid var(--accent);
      border-radius: 8px; padding: 0.95rem 1.1rem;
      font-size: 0.85rem; color: var(--text);
      margin-top: 1rem; line-height: 1.6;
    }}
    .explainer strong {{ color: var(--accent); }}
    footer {{
      text-align: center; padding: 1.5rem 1rem;
      font-size: 0.75rem; color: var(--muted-2); line-height: 1.7;
    }}
    footer a {{ color: var(--muted); }}

    /* ────────── Responsive ────────── */
    @media (max-width: 600px) {{
      .container {{ padding: 1rem; }}
      .kpi-grid {{ grid-template-columns: repeat(2, 1fr); gap: 0.6rem; }}
      .kpi-card {{ padding: 0.85rem; }}
      .kpi-value {{ font-size: 1.1rem; }}
      .header-text h1 {{ font-size: 1.05rem; }}
      .calc-results {{ grid-template-columns: 1fr; }}
      th, td {{ padding: 0.5rem 0.6rem; font-size: 0.78rem; }}
    }}
    @media (max-width: 380px) {{
      .kpi-grid {{ grid-template-columns: 1fr; }}
    }}

    /* ────────── Salary history ────────── */
    .hist-table {{ width: 100%; border-collapse: collapse; margin-bottom: 0.75rem; font-size: 0.9rem; }}
    .hist-table th {{
      text-align: left; font-size: 0.72rem; font-weight: 600; color: var(--muted);
      text-transform: uppercase; letter-spacing: 0.05em;
      padding: 0.35rem 0.5rem; border-bottom: 1px solid var(--border);
    }}
    .hist-table td {{ padding: 0.35rem 0.5rem; border-bottom: 1px solid var(--border); }}
    .hist-table td:last-child {{ width: 2.5rem; }}
    .hist-table input {{
      width: 100%; padding: 0.45rem 0.6rem; font-size: 0.9rem; font-family: inherit;
      border: 1.5px solid var(--border-strong); border-radius: 6px;
      background: var(--bg); color: var(--text);
    }}
    .hist-table input:focus {{ outline: none; border-color: var(--accent); }}
    .hist-del-btn {{
      background: none; border: 1px solid var(--border); border-radius: 6px;
      cursor: pointer; color: var(--muted); padding: 0.3rem 0.55rem;
      font-size: 1rem; line-height: 1; transition: all 0.15s ease;
    }}
    .hist-del-btn:hover {{ color: var(--negative); border-color: var(--negative); }}
    .hist-add-btn {{
      background: none; border: 1px solid var(--accent); border-radius: 8px;
      cursor: pointer; color: var(--accent); padding: 0.5rem 1rem;
      font-size: 0.85rem; font-weight: 600; font-family: inherit; transition: all 0.15s ease;
    }}
    .hist-add-btn:hover {{ background: color-mix(in srgb, var(--accent) 10%, transparent); }}
    .hist-result-tbl {{ width: 100%; border-collapse: collapse; font-size: 0.82rem; margin-top: 0.5rem; }}
    .hist-result-tbl th {{
      padding: 0.5rem 0.75rem; text-align: right; font-weight: 600; white-space: nowrap;
      font-size: 0.72rem; letter-spacing: 0.02em; text-transform: uppercase;
      background: var(--primary-dark); color: white;
    }}
    .hist-result-tbl th:first-child {{ text-align: left; }}
    .hist-result-tbl td {{
      padding: 0.5rem 0.75rem; text-align: right; border-top: 1px solid var(--border);
      white-space: nowrap; font-variant-numeric: tabular-nums;
    }}
    .hist-result-tbl td:first-child {{ text-align: left; font-weight: 600; }}
    @media (prefers-color-scheme: dark) {{
      .hist-result-tbl th {{ background: var(--border-strong); color: var(--text); }}
    }}
  </style>
</head>
<body>

<header>
  <div class="header-inner">
    <span class="header-flag">{region['flag']}</span>
    <div class="header-text">
      <h1>{region['name']} Salary Tracker</h1>
      <p>Real salary in {base_year} {region['currency']} · CPI-adjusted · updated monthly</p>
    </div>
  </div>
</header>

{nav}

<div class="container">

  <!-- ── Current salary ─────────────────────────────── -->
  <div class="section-title">Your salary &mdash; {summary['current_salary_year']}</div>
  <div class="kpi-grid">
    <div class="kpi-card">
      <div class="kpi-label">Nominal monthly</div>
      <div class="kpi-value">{fmt_cur(summary['nominal_monthly'])}</div>
      <div class="kpi-sub">Gross contractual</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Nominal annual</div>
      <div class="kpi-value">{fmt_cur(summary['nominal_annual'])}</div>
      <div class="kpi-sub">Total per year</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Nominal hourly</div>
      <div class="kpi-value">{fmt_cur(summary['nominal_hourly'], 2)}</div>
      <div class="kpi-sub">{summary['weekly_hours']:g} h/wk · {int(summary['annual_hours']):,} h/yr</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Nominal vs {base_year}</div>
      <div class="kpi-value {nom_cls}">{fmt_pct(summary['nominal_increase_vs_base_pct'])}</div>
      <div class="kpi-sub">Before inflation</div>
    </div>
  </div>

  <!-- ── Real purchasing power ──────────────────────── -->
  <div class="section-title">Real purchasing power · {base_year} {region['currency']} (CPI {summary['latest_cpi_period']})</div>
  <div class="kpi-grid">
    <div class="kpi-card">
      <div class="kpi-label">Real monthly</div>
      <div class="kpi-value">{fmt_cur(summary['real_monthly'])}</div>
      <div class="kpi-sub">In {base_year} purchasing power</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Real annual</div>
      <div class="kpi-value">{fmt_cur(summary['real_annual'])}</div>
      <div class="kpi-sub">In {base_year} purchasing power</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Real hourly</div>
      <div class="kpi-value">{fmt_cur(summary['real_hourly'], 2)}</div>
      <div class="kpi-sub">In {base_year} {region['currency']}</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Real vs {base_year}</div>
      <div class="kpi-value {real_cls}">{fmt_pct(summary['real_increase_vs_base_pct'])}</div>
      <div class="kpi-sub">After inflation</div>
    </div>
  </div>

  <!-- ── Inflation & gain/loss ──────────────────────── -->
  <div class="section-title">Inflation &amp; gain / loss vs {base_year}</div>
  <div class="kpi-grid">
    <div class="kpi-card">
      <div class="kpi-label">Cumulative inflation</div>
      <div class="kpi-value" style="color: var(--primary)">{fmt_pct(summary['cumulative_inflation_pct'])}</div>
      <div class="kpi-sub">Since {base_year} avg</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Annual gain/loss</div>
      <div class="kpi-value {gl_a_cls}">{fmt_cur(summary['annual_gain_loss_vs_base'])}</div>
      <div class="kpi-sub">Real {region['currency']} per year</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Monthly gain/loss</div>
      <div class="kpi-value {gl_m_cls}">{fmt_cur(summary['monthly_gain_loss_vs_base'])}</div>
      <div class="kpi-sub">Real {region['currency']} per month</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Latest CPI</div>
      <div class="kpi-value">{summary['latest_cpi_value']}</div>
      <div class="kpi-sub">{region['cpi_source_label']} · {summary['latest_cpi_period']}</div>
    </div>
  </div>

  <div class="explainer">
    <strong>Real salary</strong> adjusts your nominal salary for inflation, expressing it in
    {base_year} purchasing power. Formula: Real = Nominal × CPI<sub>{base_year}</sub> ÷ CPI<sub>now</sub>.
    A real increase means your salary buys <em>more</em> than in {base_year}.
  </div>

  <!-- ── Interactive calculator ─────────────────────── -->
  <div class="section-title">Try with your own data</div>
  <div class="calculator">
    <div class="calc-tabs" role="tablist">
      <button class="calc-tab active" data-tab="single" role="tab">Real salary now</button>
      <button class="calc-tab" data-tab="compare" role="tab">Compare two years</button>
      <button class="calc-tab" data-tab="history" role="tab">Salary history</button>
    </div>

    <!-- ── Mode 1: single salary ── -->
    <div class="calc-pane active" data-pane="single">
      <p class="calc-hint">Plug in your salary; see what it's worth in any reference year, with current inflation.</p>
      <div class="calc-grid">
        <div class="calc-inputs">
          <div class="calc-field">
            <label for="calc-salary">Your annual salary ({region['currency']})</label>
            <input type="number" id="calc-salary" value="{int(summary['nominal_annual'])}" min="0" step="100" inputmode="numeric">
          </div>
          <div class="calc-field">
            <label for="calc-hours">Weekly hours</label>
            <input type="number" id="calc-hours" value="{region['weekly_hours']}" min="1" step="0.5" inputmode="decimal">
          </div>
          <div class="calc-field">
            <label for="calc-base">Reference year (purchasing power baseline)</label>
            <select id="calc-base"></select>
          </div>
          <div class="calc-field">
            <label for="calc-cur">Inflation up to (CPI month)</label>
            <select id="calc-cur"></select>
          </div>
        </div>
        <div class="calc-results" id="calc-results"></div>
      </div>
    </div>

    <!-- ── Mode 2: compare two years ── -->
    <div class="calc-pane" data-pane="compare">
      <p class="calc-hint">Compare two salaries from different years. Example: <em>2010 = {region['currency_symbol']} 50,000</em> vs <em>2025 = {region['currency_symbol']} 150,000</em>. The calculator shows your <strong>real</strong> raise after inflation.</p>
      <div class="calc-grid">
        <div class="calc-inputs">
          <div class="calc-row">
            <div class="calc-field">
              <label for="cmp-year-a">From year</label>
              <select id="cmp-year-a"></select>
            </div>
            <div class="calc-field">
              <label for="cmp-sal-a">Salary then ({region['currency']})</label>
              <input type="number" id="cmp-sal-a" value="50000" min="0" step="100" inputmode="numeric">
            </div>
          </div>
          <div class="calc-row">
            <div class="calc-field">
              <label for="cmp-year-b">To year</label>
              <select id="cmp-year-b"></select>
            </div>
            <div class="calc-field">
              <label for="cmp-sal-b">Salary now ({region['currency']})</label>
              <input type="number" id="cmp-sal-b" value="80000" min="0" step="100" inputmode="numeric">
            </div>
          </div>
        </div>
        <div class="calc-results" id="cmp-results"></div>
      </div>
    </div>

    <!-- ── Mode 3: salary history ── -->
    <div class="calc-pane" data-pane="history">
      <p class="calc-hint">Enter your salary at different points in time. The calculator converts each to real purchasing power, so you can see how much your salary has actually grown after inflation.</p>
      <div style="display:flex; gap:0.75rem; margin-bottom:0.85rem; flex-wrap:wrap; align-items:flex-end;">
        <div class="calc-field" style="min-width:160px;">
          <label for="hist-base">Reference year for real values</label>
          <select id="hist-base"></select>
        </div>
      </div>
      <table class="hist-table">
        <thead>
          <tr><th>Year</th><th>Annual salary ({region['currency']})</th><th></th></tr>
        </thead>
        <tbody id="hist-rows"></tbody>
      </table>
      <button class="hist-add-btn" id="hist-add">+ Add year</button>
      <div id="hist-chart" style="margin-top:1rem; min-height:280px;"></div>
      <div id="hist-result-table" style="margin-top:0.75rem; overflow-x:auto;"></div>
    </div>
  </div>

  <!-- ── Charts ─────────────────────────────────────── -->
  <div class="section-title">Charts</div>
  <div class="chart-grid">
    <div class="chart-card"><div id="chart-yearly"></div></div>
    <div class="chart-card"><div id="chart-hourly"></div></div>
    <div class="chart-card"><div id="chart-inflation"></div></div>
    <div class="chart-card"><div id="chart-gain"></div></div>
  </div>

  <!-- ── Year-by-year table ─────────────────────────── -->
  <div class="section-title">Year-by-year detail</div>
  <div class="table-card">
    <div class="table-scroll">
      <table>
        <thead>
          <tr>
            <th>Year</th>
            <th>Nom. Monthly</th>
            <th>Nom. Annual</th>
            <th>Nom. vs {base_year}</th>
            <th>Inflation</th>
            <th>Real Monthly</th>
            <th>Real Annual</th>
            <th>Annual G/L</th>
            <th>Real vs {base_year}</th>
            <th>Nom. Hourly</th>
            <th>Real Hourly</th>
          </tr>
        </thead>
        <tbody>{table_rows}</tbody>
      </table>
    </div>
  </div>

</div>

<footer>
  Data: {region['cpi_source_label']} &middot; Generated {generated}<br>
  {region['name']} · {region['salary_note']}
</footer>

<script>
  // ────────── Charts ──────────
  var CFG = {{responsive: true, displayModeBar: false}};
  function renderChart(id, p) {{
    if (!p) return;
    var darkMode = window.matchMedia && window.matchMedia('(prefers-color-scheme: dark)').matches;
    var fontColor = darkMode ? '#cbd5e1' : '#0f172a';
    Plotly.newPlot(id, p.traces, Object.assign({{
      paper_bgcolor: "transparent",
      plot_bgcolor: "transparent",
      font: {{family: "Inter, -apple-system, sans-serif", size: 12, color: fontColor}},
    }}, p.layout), CFG);
  }}
  renderChart("chart-yearly",    {chart_yearly});
  renderChart("chart-hourly",    {chart_hourly});
  renderChart("chart-inflation", {chart_inflation});
  renderChart("chart-gain",      {chart_gain});

  // ────────── Interactive calculator ──────────
  var CALC = {calc_payload};

  // Index CPI by period
  var cpiByPeriod = {{}};
  CALC.cpi_data.forEach(function(r) {{ cpiByPeriod[r.period] = r.cpi; }});

  // Annual averages by year
  var cpiByYear = {{}};
  CALC.available_years.forEach(function(y) {{
    var vals = CALC.cpi_data.filter(function(r) {{ return r.period.indexOf(y + '-') === 0; }})
                            .map(function(r) {{ return r.cpi; }});
    if (vals.length) cpiByYear[y] = vals.reduce(function(a,b){{return a+b;}}, 0) / vals.length;
  }});

  function fmtNum(v, decimals) {{
    decimals = decimals || 0;
    if (v === null || v === undefined || isNaN(v)) return '—';
    var s = Math.abs(v).toFixed(decimals);
    var parts = s.split('.');
    parts[0] = parts[0].replace(/\\B(?=(\\d{{3}})+(?!\\d))/g, CALC.thousand_sep);
    var sign = v < 0 ? '−' : '';
    var sym = CALC.currency_symbol;
    var joined = parts.join('.');
    return (sym === '£' || sym === '$') ? sign + sym + joined : sign + sym + ' ' + joined;
  }}
  function fmtPct(v, decimals) {{
    decimals = decimals === undefined ? 1 : decimals;
    if (v === null || isNaN(v)) return '—';
    var sign = v >= 0 ? '+' : '';
    return sign + v.toFixed(decimals) + '%';
  }}

  // Populate selects
  var baseSel = document.getElementById('calc-base');
  var curSel = document.getElementById('calc-cur');
  CALC.available_years.forEach(function(y) {{
    var o = document.createElement('option');
    o.value = y; o.textContent = y;
    if (y === CALC.default_base_year) o.selected = true;
    baseSel.appendChild(o);
  }});
  CALC.cpi_data.forEach(function(r) {{
    var o = document.createElement('option');
    o.value = r.period; o.textContent = r.period;
    curSel.appendChild(o);
  }});
  curSel.selectedIndex = curSel.options.length - 1;

  function recalc() {{
    var nominal = parseFloat(document.getElementById('calc-salary').value) || 0;
    var hours = parseFloat(document.getElementById('calc-hours').value) || 0;
    var baseYear = parseInt(baseSel.value);
    var curPeriod = curSel.value;

    var baseCpi = cpiByYear[baseYear];
    var curCpi = cpiByPeriod[curPeriod];
    var annualHours = hours * CALC.default_weeks;

    var realAnnual = baseCpi && curCpi ? nominal * baseCpi / curCpi : null;
    var realMonthly = realAnnual !== null ? realAnnual / 12 : null;
    var realHourly = realAnnual !== null && annualHours > 0 ? realAnnual / annualHours : null;
    var nomHourly = annualHours > 0 ? nominal / annualHours : null;
    var nomMonthly = nominal / 12;
    var infl = baseCpi && curCpi ? (curCpi/baseCpi - 1) * 100 : null;
    var realChange = realAnnual !== null ? (realAnnual/nominal - 1) * 100 : null;
    var gainLoss = realAnnual !== null ? realAnnual - nominal : null;

    var html = '';
    function box(lbl, val, cls, full) {{
      return '<div class="calc-result' + (full ? ' full':'') + '">' +
             '<div class="lbl">' + lbl + '</div>' +
             '<div class="val' + (cls? ' ' + cls : '') + '">' + val + '</div></div>';
    }}
    html += box('Nominal monthly', fmtNum(nomMonthly));
    html += box('Nominal hourly', fmtNum(nomHourly, 2));
    html += box('Real annual (' + baseYear + ' ' + CALC.currency_symbol.replace(' ','') + ')', fmtNum(realAnnual));
    html += box('Real monthly', fmtNum(realMonthly));
    html += box('Real hourly', fmtNum(realHourly, 2));
    html += box('Cumulative inflation', fmtPct(infl), infl >= 0 ? 'negative' : 'positive');
    html += box('Real change vs nominal', fmtPct(realChange), realChange >= 0 ? 'positive' : 'negative');
    html += box('Annual gain / loss', fmtNum(gainLoss), gainLoss >= 0 ? 'positive' : 'negative', true);
    document.getElementById('calc-results').innerHTML = html;
  }}

  ['calc-salary','calc-hours'].forEach(function(id) {{
    document.getElementById(id).addEventListener('input', recalc);
  }});
  baseSel.addEventListener('change', recalc);
  curSel.addEventListener('change', recalc);
  recalc();

  // ────────── Compare two years mode ──────────
  var yearASel = document.getElementById('cmp-year-a');
  var yearBSel = document.getElementById('cmp-year-b');
  var salAEl = document.getElementById('cmp-sal-a');
  var salBEl = document.getElementById('cmp-sal-b');

  CALC.available_years.forEach(function(y) {{
    var oa = document.createElement('option');
    oa.value = y; oa.textContent = y;
    yearASel.appendChild(oa);
    var ob = document.createElement('option');
    ob.value = y; ob.textContent = y;
    yearBSel.appendChild(ob);
  }});

  // Sensible defaults: ~10 years apart, ending at most-recent year
  var lastYear = CALC.available_years[CALC.available_years.length - 1];
  var startYear = CALC.available_years.indexOf(lastYear - 10) >= 0 ? lastYear - 10 : CALC.available_years[0];
  yearASel.value = startYear;
  yearBSel.value = lastYear;

  function compare() {{
    var ya = parseInt(yearASel.value);
    var yb = parseInt(yearBSel.value);
    var sa = parseFloat(salAEl.value) || 0;
    var sb = parseFloat(salBEl.value) || 0;
    var cpiA = cpiByYear[ya];
    var cpiB = cpiByYear[yb];

    var html = '';
    function box(lbl, val, cls, headline) {{
      var c = 'calc-result';
      if (headline) c += ' full headline';
      return '<div class="' + c + '">' +
             '<div class="lbl">' + lbl + '</div>' +
             '<div class="val' + (cls? ' ' + cls : '') + '">' + val + '</div></div>';
    }}

    if (!cpiA || !cpiB || sa <= 0 || sb <= 0) {{
      document.getElementById('cmp-results').innerHTML =
        box('Need both salaries to compare', '—');
      return;
    }}

    var inflation = (cpiB / cpiA - 1) * 100;             // total price change A→B
    var nomChange = (sb / sa - 1) * 100;                 // nominal raise %
    var realB_inA = sb * cpiA / cpiB;                    // salary B in year-A money
    var realChange = (realB_inA / sa - 1) * 100;         // real raise %
    var equivToday = sa * cpiB / cpiA;                   // salary A grown to year-B money to keep parity
    var gainLoss = realB_inA - sa;                       // real CHF/£ gained vs A
    var yearsApart = yb - ya;
    var realCagr = yearsApart > 0 ? (Math.pow(realB_inA / sa, 1 / yearsApart) - 1) * 100 : 0;

    var msg;
    if (realChange > 0.5) {{
      msg = 'You got <strong>' + fmtPct(realChange) + '</strong> richer in real terms';
    }} else if (realChange < -0.5) {{
      msg = 'You got <strong>' + fmtPct(realChange) + '</strong> poorer in real terms';
    }} else {{
      msg = 'Your purchasing power is roughly the same';
    }}

    html += '<div class="calc-result full headline"><div class="lbl">Verdict (' + ya + ' → ' + yb + ')</div>' +
            '<div class="val">' + msg + '</div></div>';
    html += box('Inflation between ' + ya + ' and ' + yb, fmtPct(inflation), 'negative');
    html += box('Nominal change in salary', fmtPct(nomChange), nomChange >= 0 ? 'positive' : 'negative');
    html += box(ya + ' salary in ' + yb + ' money', fmtNum(equivToday));
    html += box(yb + ' salary in ' + ya + ' money', fmtNum(realB_inA));
    html += box('Real raise (' + yearsApart + ' yrs)', fmtPct(realChange), realChange >= 0 ? 'positive' : 'negative');
    if (yearsApart > 0) {{
      html += box('Annualised real growth', fmtPct(realCagr, 2), realCagr >= 0 ? 'positive' : 'negative');
    }}
    html += box('Real gain / loss vs ' + ya, fmtNum(gainLoss), gainLoss >= 0 ? 'positive' : 'negative', true);

    document.getElementById('cmp-results').innerHTML = html;
  }}

  [yearASel, yearBSel, salAEl, salBEl].forEach(function(el) {{
    el.addEventListener('input', compare);
    el.addEventListener('change', compare);
  }});
  compare();

  // ────────── Salary history mode ──────────
  var histBaseSel = document.getElementById('hist-base');
  CALC.available_years.forEach(function(y) {{
    var o = document.createElement('option');
    o.value = y; o.textContent = y;
    if (y === CALC.default_base_year) o.selected = true;
    histBaseSel.appendChild(o);
  }});
  histBaseSel.addEventListener('change', calcHistory);

  document.getElementById('hist-add').addEventListener('click', function() {{
    appendHistRow('', '');
    calcHistory();
  }});

  function appendHistRow(year, annual) {{
    var tbody = document.getElementById('hist-rows');
    var tr = document.createElement('tr');
    tr.innerHTML =
      '<td><input type="number" class="hist-year" value="' + (year || '') + '" min="1988" max="2030" step="1" placeholder="Year" inputmode="numeric"></td>' +
      '<td><input type="number" class="hist-sal" value="' + (annual || '') + '" min="0" step="1000" placeholder="Annual salary" inputmode="numeric"></td>' +
      '<td><button class="hist-del-btn" type="button">&times;</button></td>';
    tr.querySelector('.hist-del-btn').addEventListener('click', function() {{
      tr.parentNode.removeChild(tr);
      calcHistory();
    }});
    tr.querySelector('.hist-year').addEventListener('input', calcHistory);
    tr.querySelector('.hist-sal').addEventListener('input', calcHistory);
    tbody.appendChild(tr);
  }}

  function calcHistory() {{
    var rows = document.querySelectorAll('#hist-rows tr');
    var data = [];
    rows.forEach(function(tr) {{
      var yr = parseInt(tr.querySelector('.hist-year').value);
      var sal = parseFloat(tr.querySelector('.hist-sal').value);
      if (yr >= 1988 && yr <= 2030 && sal > 0 && cpiByYear[yr]) {{
        data.push({{year: yr, annual: sal}});
      }}
    }});
    data.sort(function(a, b) {{ return a.year - b.year; }});

    var chartEl = document.getElementById('hist-chart');
    var resultDiv = document.getElementById('hist-result-table');

    if (data.length < 2) {{
      resultDiv.innerHTML = '<p style="color:var(--muted);font-size:0.85rem;padding:0.5rem 0">Add at least 2 valid years to see results.</p>';
      if (chartEl && chartEl.data) Plotly.purge(chartEl);
      return;
    }}

    var baseYear = parseInt(histBaseSel.value);
    var baseCpi = cpiByYear[baseYear];
    var years = data.map(function(d) {{ return d.year; }});
    var nominals = data.map(function(d) {{ return d.annual; }});
    var reals = data.map(function(d) {{
      return baseCpi && cpiByYear[d.year] ? d.annual * baseCpi / cpiByYear[d.year] : null;
    }});

    var darkMode = window.matchMedia && window.matchMedia('(prefers-color-scheme: dark)').matches;
    var fontColor = darkMode ? '#cbd5e1' : '#0f172a';
    Plotly.newPlot('hist-chart', [
      {{x: years, y: nominals, name: 'Nominal', type: 'scatter', mode: 'lines+markers',
        line: {{color: '#3b82f6', width: 3}}, marker: {{size: 8}}}},
      {{x: years, y: reals, name: 'Real (' + baseYear + ')', type: 'scatter', mode: 'lines+markers',
        line: {{color: '#22c55e', width: 3}}, marker: {{size: 8}}}}
    ], {{
      paper_bgcolor: 'transparent', plot_bgcolor: 'transparent',
      font: {{family: 'Inter, -apple-system, sans-serif', size: 12, color: fontColor}},
      title: 'Salary history: nominal vs real (' + baseYear + ' ' + CALC.currency_symbol + ')',
      yaxis: {{title: CALC.currency_symbol, tickformat: ',.0f'}},
      xaxis: {{title: 'Year', type: 'category'}},
      legend: {{orientation: 'h', y: -0.2}},
      margin: {{t: 50, b: 60, l: 70, r: 20}}
    }}, {{responsive: true, displayModeBar: false}});

    var firstReal = reals[0];
    var tbl = '<table class="hist-result-tbl"><thead><tr>' +
      '<th>Year</th><th>Nominal</th><th>Real (' + baseYear + ')</th>' +
      '<th>Inflation vs ' + baseYear + '</th><th>Real vs ' + data[0].year + '</th>' +
      '</tr></thead><tbody>';
    data.forEach(function(d, i) {{
      var r = reals[i];
      var infl = baseCpi && cpiByYear[d.year] ? (cpiByYear[d.year] / baseCpi - 1) * 100 : null;
      var realChg = r !== null && firstReal !== null && firstReal !== 0 ? (r / firstReal - 1) * 100 : null;
      var chgCls = realChg === null ? '' : (realChg >= 0 ? ' class="positive"' : ' class="negative"');
      tbl += '<tr><td>' + d.year + '</td><td>' + fmtNum(d.annual) + '</td><td>' + fmtNum(r) +
             '</td><td>' + fmtPct(infl) + '</td><td' + chgCls + '>' + fmtPct(realChg) + '</td></tr>';
    }});
    tbl += '</tbody></table>';
    resultDiv.innerHTML = tbl;
  }}

  CALC.salary_history_defaults.forEach(function(d) {{
    appendHistRow(d.year, d.annual);
  }});
  calcHistory();

  // ────────── Tabs ──────────
  document.querySelectorAll('.calc-tab').forEach(function(btn) {{
    btn.addEventListener('click', function() {{
      var tab = btn.dataset.tab;
      document.querySelectorAll('.calc-tab').forEach(function(b) {{ b.classList.toggle('active', b === btn); }});
      document.querySelectorAll('.calc-pane').forEach(function(p) {{ p.classList.toggle('active', p.dataset.pane === tab); }});
      if (tab === 'history') {{
        var hc = document.getElementById('hist-chart');
        if (hc && hc.data && hc.data.length) Plotly.relayout('hist-chart', {{autosize: true}});
      }}
    }});
  }});
</script>
</body>
</html>
"""
    return html


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def build_excel(
    summary: dict,
    df_yearly: pd.DataFrame,
    df_monthly: pd.DataFrame,
    salary_path: Path,
    cpi_path: Path,
    output_path: Path,
) -> None:
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        log.warning("openpyxl not installed; skipping Excel output.")
        return

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Salary_Input"
    df_sal = pd.read_csv(salary_path)
    for r in dataframe_to_rows(df_sal, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("CPI_Data")
    if cpi_path.exists():
        df_cpi_raw = pd.read_csv(cpi_path)
        for r in dataframe_to_rows(df_cpi_raw, index=False, header=True):
            ws2.append(r)

    ws3 = wb.create_sheet("Calculations")
    for r in dataframe_to_rows(df_yearly.round(2), index=False, header=True):
        ws3.append(r)

    ws4 = wb.create_sheet("Summary")
    ws4.append(["Metric", "Value"])
    for k, v in summary.items():
        ws4.append([k, v])

    wb.save(output_path)
    log.info("Excel saved → %s", output_path)


# ---------------------------------------------------------------------------
# Entry: build a single region's HTML page
# ---------------------------------------------------------------------------

def build(
    summary: dict,
    df_yearly: pd.DataFrame,
    df_monthly: pd.DataFrame,
    df_cpi_full: pd.DataFrame,
    region: dict,
    all_regions: Iterable[dict],
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    html = build_html(summary, df_yearly, df_monthly, df_cpi_full, region, all_regions)
    output_path.write_text(html, encoding="utf-8")
    log.info("Dashboard written → %s", output_path)
